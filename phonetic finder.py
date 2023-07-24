import time
import requests
import openpyxl
from bs4 import BeautifulSoup

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36'
}

def get_phonetics(word):
    url = f"https://dictionary.cambridge.org/dictionary/english/{word}"
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.text, 'html.parser')
    
    try:
        uk_phonetic = soup.find('span', class_='uk dpron-i').find('span', class_='ipa dipa lpr-2 lpl-1').text
    except AttributeError:
        uk_phonetic = "Not found"
    
    try:
        us_phonetic = soup.find('span', class_='us dpron-i').find('span', class_='ipa dipa lpr-2 lpl-1').text
    except AttributeError:
        us_phonetic = "Not found"
    
    return uk_phonetic, us_phonetic


def update_last_row(row):
    with open('last_row.txt', 'w') as f:
        f.write(str(row))

def get_last_row():
    try:
        with open('last_row.txt', 'r') as f:
            return int(f.read().strip())
    except Exception as e:
        return 2  # default value if file doesn't exist

# Load workbook and sheet
wb = openpyxl.load_workbook('words.xlsx')
ws = wb.active

start_row = get_last_row()
n = start_row - 2

for row in range(start_row, ws.max_row + 1):
    n += 1
    word = ws.cell(row=row, column=2).value  # Get word
    uk_phonetic, us_phonetic = get_phonetics(word)  # Get phonetics
    ws.cell(row=row, column=3).value = us_phonetic  # Write US phonetics
    ws.cell(row=row, column=4).value = uk_phonetic  # Write UK phonetics
    
    time.sleep(2)
    
    print('counting=',n)
    if n==972:
        break
    else:
        # Save workbook and last completed row
        wb.save('words.xlsx')
        update_last_row(row)

# Save workbook one last time to ensure all changes are saved
wb.save('words.xlsx')
